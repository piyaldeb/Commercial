import requests
import json
import base64
import logging
import sys
import os

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr.encoding and sys.stderr.encoding.lower() != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8')

from datetime import date, datetime, timedelta
import gspread
from gspread_dataframe import set_with_dataframe
from google.oauth2 import service_account
import pandas as pd
import time
from dotenv import load_dotenv
from requests.exceptions import RequestException

load_dotenv()
logging.basicConfig(stream=sys.stdout, level=logging.INFO)
log = logging.getLogger()

# ========= CONFIG ==========
ODOO_URL = os.getenv("ODOO_URL")
DB = os.getenv("ODOO_DB")
USERNAME = os.getenv("ODOO_USERNAME")
PASSWORD = os.getenv("ODOO_PASSWORD")

SHEET_KEY = "1coN06mZ9uLBn1JnSyNqLwYhpl2-fJsuT85xwuNI0iy8"
WORKSHEET_NAME = "Raw"
DUMP_WORKSHEET_NAME = "Dump"
DUMP_MONTHS = 3

today = date.today()

session = requests.Session()
USER_ID = None

# ========= GOOGLE SHEETS CLIENT ==========
def get_gspread_client():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for sa_path in [
        os.path.join(script_dir, "service_account.json"),
        "service_account.json",
    ]:
        if os.path.exists(sa_path):
            return gspread.service_account(filename=sa_path)

    # GitHub Actions style: full JSON secret stored in GOOGLE_SHEET_CRED_JSON
    creds_json = os.getenv("GOOGLE_SHEET_CRED_JSON")
    if creds_json:
        creds_dict = json.loads(creds_json)
        scopes = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = service_account.Credentials.from_service_account_info(creds_dict, scopes=scopes)
        return gspread.authorize(creds)

    creds_raw = os.getenv("GOOGLE_CREDS_BASE64")
    if creds_raw:
        creds_dict = None
        try:
            creds_dict = json.loads(creds_raw.strip())
        except json.JSONDecodeError:
            pass
        if creds_dict is None:
            try:
                padded = creds_raw.strip() + '=' * (-len(creds_raw.strip()) % 4)
                creds_dict = json.loads(base64.b64decode(padded).decode("utf-8"))
            except Exception:
                pass
        if creds_dict is None:
            raise Exception("GOOGLE_CREDS_BASE64 is neither valid JSON nor valid base64-encoded JSON")
        scopes = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = service_account.Credentials.from_service_account_info(creds_dict, scopes=scopes)
        return gspread.authorize(creds)

    raise Exception("No Google credentials found.")

# ========= RETRY LOGIC ==========
def retry_request(method, url, max_retries=3, backoff=3, **kwargs):
    for attempt in range(1, max_retries + 1):
        try:
            r = method(url, **kwargs)
            r.raise_for_status()
            return r
        except RequestException as e:
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt < max_retries:
                print(f"⏳ Retrying in {backoff} seconds...")
                time.sleep(backoff)
            else:
                print("❌ All retry attempts failed.")
                raise

# ========= LOGIN ==========
def login():
    global USER_ID
    payload = {"jsonrpc": "2.0", "params": {"db": DB, "login": USERNAME, "password": PASSWORD}}
    r = retry_request(session.post, f"{ODOO_URL}/web/session/authenticate", json=payload)
    result = r.json().get("result")
    if result and "uid" in result:
        USER_ID = result["uid"]
        print(f"✅ Logged in (uid={USER_ID})")
        return result
    else:
        raise Exception("❌ Login failed")

# ========= SWITCH COMPANY ==========
def switch_company(company_id):
    if USER_ID is None:
        raise Exception("User not logged in yet")
    payload = {
        "jsonrpc": "2.0",
        "method": "call",
        "params": {
            "model": "res.users",
            "method": "write",
            "args": [[USER_ID], {"company_id": company_id}],
            "kwargs": {"context": {"allowed_company_ids": [company_id], "company_id": company_id}},
        },
    }
    r = retry_request(session.post, f"{ODOO_URL}/web/dataset/call_kw", json=payload)
    if "error" in r.json():
        print(f"❌ Failed to switch to company {company_id}: {r.json()['error']}")
        return False
    print(f"🔄 Session switched to company {company_id}")
    return True

# ========= FETCH COMBINE INVOICES ==========
def fetch_combine_invoices(allowed_company_ids, invoice_date_from=None):
    ctx = {
        "lang": "en_US", "tz": "Asia/Dhaka", "uid": USER_ID,
        "allowed_company_ids": allowed_company_ids, "bin_size": True,
        "current_company_id": allowed_company_ids[0],
    }
    spec = {
        "name":                    {},
        "create_date":             {},
        "report_date":             {},
        "delivery_date":           {},
        "create_uid":              {"fields": {"display_name": {}}},
        "partner_id":              {"fields": {"display_name": {}}},
        "invoice_date":            {},
        "invoice_payment_term_id": {"fields": {"display_name": {}}},
        "invoice_incoterm_id":     {"fields": {"display_name": {}}},
        "sales_person":            {"fields": {"display_name": {}}},
        "team_id":                 {"fields": {"display_name": {}}},
        "buyer_name":              {"fields": {"display_name": {}}},
        "z_total_q":               {},
        "m_total_q":               {},
        "z_total":                 {},
        "m_total":                 {},
        "qty_total":               {},
        "amount_total":            {},
        "state":                   {},
    }
    domain = [["state", "=", "posted"]]
    if invoice_date_from:
        domain = ["&", ["state", "=", "posted"], ["invoice_date", ">=", invoice_date_from]]

    all_records = []
    offset = 0
    page = 5000
    while True:
        payload = {
            "jsonrpc": "2.0",
            "method": "call",
            "params": {
                "model": "combine.invoice",
                "method": "web_search_read",
                "args": [],
                "kwargs": {
                    "specification": spec,
                    "offset": offset,
                    "order": "invoice_date DESC",
                    "limit": page,
                    "context": ctx,
                    "count_limit": 100001,
                    "domain": domain,
                },
            },
        }
        r = retry_request(
            session.post,
            f"{ODOO_URL}/web/dataset/call_kw/combine.invoice/web_search_read",
            json=payload,
        )
        body = r.json()
        if "error" in body:
            raise Exception(f"combine.invoice fetch failed: {body['error'].get('message') or body['error']}")
        recs = body.get("result", {}).get("records", [])
        all_records.extend(recs)
        print(f"📋 combine.invoice: fetched {len(recs)} (total {len(all_records)})")
        if len(recs) < page:
            break
        offset += page
    return all_records

# ========= FETCH COMBINE INVOICE LINES (per invoice → sale_order_line) ==========
def fetch_combine_invoice_lines(allowed_company_ids, invoice_ids):
    """Return list of {invoice_id, sale_order_line_id, delivery_date} for given invoice IDs."""
    if not invoice_ids:
        return []
    ctx = {
        "lang": "en_US", "tz": "Asia/Dhaka", "uid": USER_ID,
        "allowed_company_ids": allowed_company_ids, "bin_size": True,
        "current_company_id": allowed_company_ids[0],
    }
    spec = {
        "invoice_id":      {"fields": {"display_name": {}}},
        "sale_order_line": {"fields": {"display_name": {}}},
        "delivery_date":   {},
        "quantity":        {},
        "price_subtotal":  {},
    }
    out = []
    CHUNK = 500
    for i in range(0, len(invoice_ids), CHUNK):
        chunk = invoice_ids[i:i + CHUNK]
        offset = 0
        page = 5000
        while True:
            payload = {
                "jsonrpc": "2.0",
                "method": "call",
                "params": {
                    "model": "combine.invoice.line",
                    "method": "web_search_read",
                    "args": [],
                    "kwargs": {
                        "specification": spec,
                        "offset": offset,
                        "order": "id ASC",
                        "limit": page,
                        "context": ctx,
                        "count_limit": 100001,
                        "domain": [["invoice_id", "in", chunk]],
                    },
                },
            }
            r = retry_request(
                session.post,
                f"{ODOO_URL}/web/dataset/call_kw/combine.invoice.line/web_search_read",
                json=payload,
            )
            body = r.json()
            if "error" in body:
                raise Exception(f"combine.invoice.line fetch failed: {body['error'].get('message') or body['error']}")
            recs = body.get("result", {}).get("records", [])
            out.extend(recs)
            if len(recs) < page:
                break
            offset += page
    print(f"📋 combine.invoice.line: fetched {len(out)}")
    return out


# ========= FETCH OPERATION DETAILS (Delivery, by sale_order_line) ==========
def fetch_operation_details_by_sale_lines(allowed_company_ids, sale_line_ids,
                                          action_date_from=None, action_date_to=None):
    """Fetch operation.details with next_operation=Delivery for the given sale_order_line IDs."""
    if not sale_line_ids:
        return []
    ctx = {
        "lang": "en_US", "tz": "Asia/Dhaka", "uid": USER_ID,
        "allowed_company_ids": allowed_company_ids, "bin_size": True,
        "current_company_id": allowed_company_ids[0],
    }
    spec = {
        "sale_order_line":  {"fields": {"display_name": {}}},
        "oa_id":            {"fields": {"display_name": {}}},
        "action_date":      {},
        "action_only_date": {},
        "qty":              {},
        "final_price":      {},
        "next_operation":   {},
        "state":            {},
    }

    out = []
    CHUNK = 500
    ids_list = list(set(sale_line_ids))
    for i in range(0, len(ids_list), CHUNK):
        chunk = ids_list[i:i + CHUNK]
        domain = [
            "&", "&",
            ["sale_order_line", "in", chunk],
            ["next_operation", "=", "Delivery"],
            ["state", "!=", "done"],
        ]
        if action_date_from and action_date_to:
            domain = [
                "&", "&", "&", "&",
                ["sale_order_line", "in", chunk],
                ["next_operation", "=", "Delivery"],
                ["state", "!=", "done"],
                ["action_date", ">=", action_date_from],
                ["action_date", "<", action_date_to],
            ]
        elif action_date_from:
            domain = [
                "&", "&", "&",
                ["sale_order_line", "in", chunk],
                ["next_operation", "=", "Delivery"],
                ["state", "!=", "done"],
                ["action_date", ">=", action_date_from],
            ]
        offset = 0
        page = 5000
        while True:
            payload = {
                "jsonrpc": "2.0",
                "method": "call",
                "params": {
                    "model": "operation.details",
                    "method": "web_search_read",
                    "args": [],
                    "kwargs": {
                        "specification": spec,
                        "offset": offset,
                        "order": "action_date ASC",
                        "limit": page,
                        "context": ctx,
                        "count_limit": 100001,
                        "domain": domain,
                    },
                },
            }
            r = retry_request(
                session.post,
                f"{ODOO_URL}/web/dataset/call_kw/operation.details/web_search_read",
                json=payload,
            )
            body = r.json()
            if "error" in body:
                raise Exception(f"operation.details fetch failed: {body['error'].get('message') or body['error']}")
            recs = body.get("result", {}).get("records", [])
            out.extend(recs)
            if len(recs) < page:
                break
            offset += page
    print(f"📋 operation.details (by sale_order_line): fetched {len(out)}")
    return out

# ========= BUILD REPORT ==========
def build_report(invoices, invoice_lines, ops, month_start):
    """
    Join:
      combine.invoice  --(line_id)-->  combine.invoice.line  --(sale_order_line)-->  operation.details
    For each invoice, sum qty * final_price from operation.details (Delivery lines)
    matching that invoice's sale_order_line IDs, bucketed by action_only_date.

    Returns (rows, weeks). rows have keys:
        Invoice No, Invoice Date, Customer, Full Qty Delivery Date.,
        Invoice Value, Invoice QTY, _day_values (dict: date -> value)
    """
    import calendar as _cal

    year, month = month_start.year, month_start.month
    last_day = _cal.monthrange(year, month)[1]

    # Sun-Thu workweeks fully inside the month
    weeks = []
    d = date(year, month, 1)
    while d.weekday() != 6 and d.month == month:
        d += timedelta(days=1)
    while d.month == month:
        wk = [d + timedelta(days=i) for i in range(5)]
        if wk[-1].month == month and wk[-1].day <= last_day:
            weeks.append(wk)
        d += timedelta(days=7)
    all_days = set(day for wk in weeks for day in wk)

    def _to_day(v):
        try:
            return pd.to_datetime(v).date() if v else None
        except Exception:
            return None

    # ---- invoice_id → list of sale_order_line IDs ----
    inv_to_sale_lines = {}
    for il in invoice_lines:
        inv_id = (il.get("invoice_id") or {}).get("id")
        sol = (il.get("sale_order_line") or {}).get("id")
        if inv_id is None or sol is None:
            continue
        inv_to_sale_lines.setdefault(inv_id, set()).add(sol)

    # ---- (sale_order_line, day) → summed qty*final_price ----
    op_lookup = {}
    for op in ops:
        sol = (op.get("sale_order_line") or {}).get("id")
        if sol is None:
            continue
        # Prefer action_only_date; fall back to date(action_date) when it's missing.
        # Ref behavior: lines with action_only_date=False but action_date set ARE included.
        day = _to_day(op.get("action_only_date") or op.get("action_date"))
        if day is None:
            continue
        qty = op.get("qty") or 0
        fp  = op.get("final_price") or 0
        try:
            v = float(qty) * float(fp)
        except Exception:
            v = 0.0
        if not v:
            continue
        op_lookup[(sol, day)] = op_lookup.get((sol, day), 0.0) + v

    # ---- per-invoice daily values ----
    rows = []
    for inv in invoices:
        inv_id = inv.get("id")
        ddate = _to_day(inv.get("delivery_date"))
        day_values = {}
        sol_ids = inv_to_sale_lines.get(inv_id, set())
        for (sol, day), v in op_lookup.items():
            if sol in sol_ids and day in all_days:
                day_values[day] = day_values.get(day, 0.0) + v
        rows.append({
            "Invoice No":              inv.get("name") or "",
            "Invoice Date":            _to_day(inv.get("invoice_date")),
            "Customer":                (inv.get("partner_id") or {}).get("display_name", ""),
            "Full Qty Delivery Date.": ddate,
            "Invoice Value":           inv.get("amount_total"),
            "Invoice QTY":             inv.get("qty_total"),
            "_day_values":             day_values,
        })
    return rows, weeks


# ========= WRITE EXCEL (exact layout of "Partial Deliveries Monthly.xlsx") =========
def write_partial_deliveries_xlsx(rows, weeks, month_start, out_path):
    """
    Write a workbook with one sheet named like 'May26', laid out per the reference file:
      row 3: merged "Week 01"..."Week NN" labels above each week's day columns
      row 4: headers — Invoice No, Invoice Date, Customer, Full Qty Dalivery Date.,
             Partial Delivery total value, Invoice Value, Invoice QTY,
             then per week: <day1>..<day5>, Total Week NN
      row 5+: invoice rows; column E and Total Week NN are formulas.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = month_start.strftime("%b%y")  # e.g. "May26"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    week_fill = PatternFill("solid", fgColor="BDD7EE")
    total_fill = PatternFill("solid", fgColor="FCE4D6")

    # ---- column plan ----
    front = ["Invoice No", "Invoice Date", "Customer", "Full Qty Dalivery Date.",
            "Partial Delivery total value", "Invoice Value", "Invoice QTY"]
    n_front = len(front)  # 7 → cols A..G

    # day_layout: list of (header_label, kind, week_idx, day_obj_or_None)
    # kind: "day" or "total"
    layout = list(enumerate(front, start=1))  # (col_index, header)
    col = n_front + 1
    week_day_cols = []   # list of (week_idx, [col indices for days])
    week_total_cols = [] # list of (week_idx, total col index)
    for wi, wk in enumerate(weeks, start=1):
        day_cols_for_week = []
        for day in wk:
            layout.append((col, day.strftime("%d-%m-%Y")))
            day_cols_for_week.append((col, day))
            col += 1
        week_day_cols.append((wi, day_cols_for_week))
        layout.append((col, f"Total Week {wi:02d}"))
        week_total_cols.append((wi, col))
        col += 1

    last_col = col - 1

    # ---- row 3: merged week-band labels ----
    for (wi, day_cols_for_week), (_, total_col) in zip(week_day_cols, week_total_cols):
        first_day_col = day_cols_for_week[0][0]
        # Merge from first day column through Total column (matches the reference file)
        ws.merge_cells(start_row=3, start_column=first_day_col, end_row=3, end_column=total_col)
        c = ws.cell(row=3, column=first_day_col, value=f"Week {wi:02d}")
        c.font = bold
        c.alignment = center
        c.fill = week_fill
        c.border = border

    # ---- row 4: headers ----
    for col_idx, header in layout:
        c = ws.cell(row=4, column=col_idx, value=header)
        c.font = bold
        c.alignment = center
        c.fill = header_fill
        c.border = border

    # ---- data rows ----
    partial_total_col_letter = get_column_letter(5)  # col E
    inv_date_col = 2
    delivery_col = 4

    for ri, row in enumerate(rows, start=5):
        ws.cell(row=ri, column=1, value=row["Invoice No"]).border = border
        ws.cell(row=ri, column=2, value=row["Invoice Date"]).border = border
        ws.cell(row=ri, column=3, value=row["Customer"]).border = border
        ws.cell(row=ri, column=4, value=row["Full Qty Delivery Date."]).border = border
        # Partial Delivery total value (formula: sum of Total Week NN cells)
        if week_total_cols:
            parts = [f"{get_column_letter(c)}{ri}" for _, c in week_total_cols]
            ws.cell(row=ri, column=5, value=f"={'+'.join(parts)}").border = border
        ws.cell(row=ri, column=6, value=row["Invoice Value"]).border = border
        ws.cell(row=ri, column=7, value=row["Invoice QTY"]).border = border

        # per-day values
        for (wi, day_cols_for_week), (_, total_col) in zip(week_day_cols, week_total_cols):
            for cidx, day in day_cols_for_week:
                v = row["_day_values"].get(day)
                ws.cell(row=ri, column=cidx, value=v).border = border
            first = get_column_letter(day_cols_for_week[0][0])
            last = get_column_letter(day_cols_for_week[-1][0])
            cell = ws.cell(row=ri, column=total_col, value=f"=SUM({first}{ri}:{last}{ri})")
            cell.fill = total_fill
            cell.border = border

        # number / date formatting
        for c_idx in (2, 4):
            ws.cell(row=ri, column=c_idx).number_format = "dd-mm-yyyy"
        for c_idx in [5, 6, 7] + [c for _, c in week_total_cols] + [c for _, daycols in week_day_cols for c, _ in daycols]:
            cell = ws.cell(row=ri, column=c_idx)
            if cell.value is not None:
                cell.number_format = "#,##0.00"

    # ---- column widths ----
    widths = {1: 22, 2: 12, 3: 36, 4: 18, 5: 18, 6: 14, 7: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for _, daycols in week_day_cols:
        for c, _ in daycols:
            ws.column_dimensions[get_column_letter(c)].width = 12
    for _, c in week_total_cols:
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = "H5"

    wb.save(out_path)
    return out_path

# ========= LAST-N-MONTHS HELPERS =========
def last_n_month_starts(anchor, n):
    """Return [oldest, ..., anchor's month start] for the last n calendar months."""
    starts = []
    y, m = anchor.year, anchor.month
    for _ in range(n):
        starts.append(date(y, m, 1))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(starts))


def build_dump_rows(invoices, invoice_lines, ops, month_starts):
    """
    Produce long-format rows for the Dump sheet across multiple months.

    Columns:
        Month, Invoice No, Invoice Date, Customer, Full Qty Delivery Date,
        Invoice Value, Invoice QTY, Date, Value
    One output row per (invoice, day) that has a non-zero partial-delivery value
    within that month's Sun-Thu workdays.
    """
    out = []
    for ms in month_starts:
        rows, weeks = build_report(invoices, invoice_lines, ops, ms)
        all_days = {d for wk in weeks for d in wk}
        month_label = ms.strftime("%b%y")  # e.g. "May26"
        for r in rows:
            for d, v in r["_day_values"].items():
                if d not in all_days or not v:
                    continue
                out.append({
                    "Month":                  month_label,
                    "Invoice No":             r["Invoice No"],
                    "Invoice Date":           r["Invoice Date"].isoformat() if r["Invoice Date"] else "",
                    "Customer":               r["Customer"],
                    "Full Qty Delivery Date": r["Full Qty Delivery Date."].isoformat() if r["Full Qty Delivery Date."] else "",
                    "Invoice Value":          r["Invoice Value"] if r["Invoice Value"] is not None else "",
                    "Invoice QTY":            r["Invoice QTY"] if r["Invoice QTY"] is not None else "",
                    "Date":                   d.isoformat(),
                    "Value":                  v,
                })
    return out


# ========= MAIN ==========
if __name__ == "__main__":
    userinfo = login()
    print("User info (allowed companies):", userinfo.get("user_companies", {}))

    # company set used in the captured HARs: [3, 2, 1]; active company = 3
    company_id = 3
    allowed = [3, 2, 1]

    if not switch_company(company_id):
        sys.exit(1)

    # Fetch invoices covering the last DUMP_MONTHS calendar months so the Dump
    # sheet has data for the full window. The per-month filter still happens
    # later in build_report (Sun-Thu workdays of each month).
    month_starts_window = last_n_month_starts(today, DUMP_MONTHS)
    REPORT_FROM = month_starts_window[0]

    invoices = fetch_combine_invoices(allowed, invoice_date_from=REPORT_FROM.isoformat())
    print(f"✅ {len(invoices)} combine invoices fetched (since {REPORT_FROM})")

    # combine.invoice.line for those invoices → gives us sale_order_line IDs to join on
    invoice_ids = [i["id"] for i in invoices if i.get("id")]
    invoice_lines = fetch_combine_invoice_lines(allowed, invoice_ids)
    print(f"✅ {len(invoice_lines)} combine invoice lines fetched")

    sol_ids = sorted({(il.get("sale_order_line") or {}).get("id")
                      for il in invoice_lines
                      if (il.get("sale_order_line") or {}).get("id")})
    print(f"   {len(sol_ids)} unique sale_order_line IDs to fetch deliveries for")

    # operation.details (Delivery lines) for those sale_order_line IDs
    af = (datetime.combine(REPORT_FROM, datetime.min.time()) - timedelta(hours=6)).strftime("%Y-%m-%d %H:%M:%S")
    ops = fetch_operation_details_by_sale_lines(allowed, sol_ids, action_date_from=af)
    print(f"✅ {len(ops)} delivery operation lines fetched")

    # build for current calendar month
    month_start = date(today.year, today.month, 1)

    rows, weeks = build_report(invoices, invoice_lines, ops, month_start)

    # Match ref behavior: include an invoice if it has any partial-delivery
    # activity in the month's workdays (drives the non-zero rows in the sheet).
    all_days = {d for wk in weeks for d in wk}
    rows = [r for r in rows
            if any(d in all_days for d in r["_day_values"].keys())]
    print(f"📊 Final report: {len(rows)} rows, {len(weeks)} weeks")

    output_file = f"Partial Deliveries Monthly - {month_start.strftime('%b%y')}.xlsx"
    write_partial_deliveries_xlsx(rows, weeks, month_start, output_file)
    print(f"📂 Saved: {output_file}")

    # Google Sheets push — plain values version (resolve formulas to numbers)
    try:
        df_flat_cols = ["Invoice No", "Invoice Date", "Customer", "Full Qty Dalivery Date.",
                        "Partial Delivery total value", "Invoice Value", "Invoice QTY"]
        for wi, wk in enumerate(weeks, start=1):
            for d in wk:
                df_flat_cols.append(d.strftime("%d-%m-%Y"))
            df_flat_cols.append(f"Total Week {wi:02d}")

        flat_rows = []
        for r in rows:
            week_totals = []
            day_cells = []
            for wk in weeks:
                wt = 0.0
                for d in wk:
                    v = r["_day_values"].get(d)
                    day_cells.append(v if v is not None else "")
                    if isinstance(v, (int, float)):
                        wt += v
                week_totals.append(wt)
            grand_total = sum(week_totals)
            base = [
                r["Invoice No"],
                r["Invoice Date"].isoformat() if r["Invoice Date"] else "",
                r["Customer"],
                r["Full Qty Delivery Date."].isoformat() if r["Full Qty Delivery Date."] else "",
                grand_total,
                r["Invoice Value"] if r["Invoice Value"] is not None else "",
                r["Invoice QTY"]   if r["Invoice QTY"]   is not None else "",
            ]
            # interleave day cells with weekly totals at the right positions
            interleaved = []
            day_iter = iter(day_cells)
            for wi, wk in enumerate(weeks):
                for _ in wk:
                    interleaved.append(next(day_iter))
                interleaved.append(week_totals[wi])
            flat_rows.append(base + interleaved)

        df_flat = pd.DataFrame(flat_rows, columns=df_flat_cols)

        client = get_gspread_client()
        sheet = client.open_by_key(SHEET_KEY)
        try:
            worksheet = sheet.worksheet(WORKSHEET_NAME)
        except gspread.WorksheetNotFound:
            worksheet = sheet.add_worksheet(title=WORKSHEET_NAME,
                                            rows=max(1000, len(df_flat) + 50),
                                            cols=max(40, len(df_flat.columns) + 5))
        worksheet.clear()
        set_with_dataframe(worksheet, df_flat)
        print(f"✅ Data pasted to Google Sheets → '{WORKSHEET_NAME}'")

        # ----- Dump sheet: long-format rolling N months -----
        dump_rows = build_dump_rows(invoices, invoice_lines, ops, month_starts_window)
        dump_cols = ["Month", "Invoice No", "Invoice Date", "Customer",
                     "Full Qty Delivery Date", "Invoice Value", "Invoice QTY",
                     "Date", "Value"]
        df_dump = pd.DataFrame(dump_rows, columns=dump_cols)
        try:
            dump_ws = sheet.worksheet(DUMP_WORKSHEET_NAME)
        except gspread.WorksheetNotFound:
            dump_ws = sheet.add_worksheet(title=DUMP_WORKSHEET_NAME,
                                          rows=max(1000, len(df_dump) + 50),
                                          cols=max(15, len(df_dump.columns) + 2))
        dump_ws.clear()
        set_with_dataframe(dump_ws, df_dump)
        months_str = ", ".join(ms.strftime("%b%y") for ms in month_starts_window)
        print(f"✅ Dump pasted to Google Sheets → '{DUMP_WORKSHEET_NAME}' ({len(df_dump)} rows; months: {months_str})")
    except Exception as e:
        import traceback
        print(f"❌ Error while pasting to Google Sheets: {e}")
        traceback.print_exc()
